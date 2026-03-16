from __future__ import annotations

import os
import re
import smtplib
import ssl
import tempfile
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import joinedload

from app.core.database import SessionLocal
from .models import (
    QualityDocument,
    QualityDocumentCategory,
    QualityDocumentNotificationLog,
    QualityDocumentRevision,
)

EXCEL_HEADERS = [
    "DOK. NO",
    "DOKÜMAN ADI",
    "İLK YAY. TAR.",
    "REV. NO",
    "REV. TAR",
    "SON KONTROL TARİHİ",
    "DOKÜMAN DAĞITIM PLANI",
    "AÇIKLAMALAR",
]


def _parse_excel_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value or value == "-":
            return None
        for fmt in ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_iso(value: Optional[datetime]) -> Optional[str]:
    return value.isoformat() if value else None


def _serialize_revision_entity(item: QualityDocumentRevision):
    return {
        "id": item.id,
        "document_id": item.document_id,
        "revision_no": item.revision_no,
        "revision_date": _to_iso(item.revision_date),
        "effective_date": _to_iso(item.effective_date),
        "last_review_date": _to_iso(item.last_review_date),
        "change_summary": item.change_summary,
        "notes": item.notes,
        "distribution_text": item.distribution_text,
        "file_path": item.file_path,
        "status": item.status,
        "published_by": item.published_by,
        "published_at": _to_iso(item.published_at),
        "created_at": _to_iso(item.created_at),
        "updated_at": _to_iso(item.updated_at),
    }


def _display_date(value: Optional[datetime]) -> str:
    if not value:
        return "-"
    return value.strftime("%d.%m.%Y")


def _email_candidates_from_distribution_text(distribution_text: str) -> list[str]:
    text = _normalize_text(distribution_text)
    if not text:
        return []
    matches = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    unique: list[str] = []
    for item in matches:
        lowered = item.lower()
        if lowered not in unique:
            unique.append(lowered)
    return unique


def _fallback_notification_recipients() -> list[str]:
    raw = os.getenv("QUALITY_NOTIFICATION_EMAILS", "")
    return [item.strip() for item in raw.split(",") if item.strip()]


def _smtp_config() -> dict[str, Any]:
    return {
        "host": os.getenv("QUALITY_SMTP_HOST", "").strip(),
        "port": int(os.getenv("QUALITY_SMTP_PORT", "587") or "587"),
        "username": os.getenv("QUALITY_SMTP_USERNAME", "").strip(),
        "password": os.getenv("QUALITY_SMTP_PASSWORD", "").strip(),
        "sender": os.getenv("QUALITY_SMTP_SENDER", "").strip(),
        "use_tls": os.getenv("QUALITY_SMTP_USE_TLS", "1").strip() != "0",
        "enabled": os.getenv("QUALITY_SMTP_ENABLED", "0").strip() == "1",
    }


def list_document_categories(active_only: bool = False):
    db = SessionLocal()
    try:
        query = db.query(QualityDocumentCategory)
        if active_only:
            query = query.filter(QualityDocumentCategory.is_active == True)
        return query.order_by(QualityDocumentCategory.sort_order.asc(), QualityDocumentCategory.id.asc()).all()
    finally:
        db.close()


def create_document_category(code: str, name: str, sort_order: int = 0):
    db = SessionLocal()
    try:
        item = QualityDocumentCategory(
            code=code.strip(),
            name=name.strip(),
            sort_order=sort_order,
            is_active=True,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def get_or_create_document_category(code: str, name: Optional[str] = None, sort_order: int = 0):
    db = SessionLocal()
    try:
        item = db.query(QualityDocumentCategory).filter(QualityDocumentCategory.code == code.strip()).first()
        if item:
            return item
        item = QualityDocumentCategory(
            code=code.strip(),
            name=(name or code).strip(),
            sort_order=sort_order,
            is_active=True,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def list_documents(category_id: Optional[int] = None, active_only: bool = False):
    db = SessionLocal()
    try:
        query = db.query(QualityDocument).options(joinedload(QualityDocument.category))
        if category_id:
            query = query.filter(QualityDocument.category_id == category_id)
        if active_only:
            query = query.filter(QualityDocument.is_active == True)
        return query.order_by(QualityDocument.document_no.asc(), QualityDocument.id.asc()).all()
    finally:
        db.close()


def create_document(
    category_id: int,
    document_no: str,
    document_name: str,
    document_type: Optional[str] = None,
    department: Optional[str] = None,
    first_publish_date: Optional[datetime] = None,
):
    db = SessionLocal()
    try:
        item = QualityDocument(
            category_id=category_id,
            document_no=document_no.strip(),
            document_name=document_name.strip(),
            document_type=(document_type or "").strip() or None,
            department=(department or "").strip() or None,
            first_publish_date=first_publish_date,
            is_active=True,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def get_or_create_document(
    category_id: int,
    document_no: str,
    document_name: str,
    document_type: Optional[str] = None,
    department: Optional[str] = None,
    first_publish_date: Optional[datetime] = None,
):
    db = SessionLocal()
    try:
        item = db.query(QualityDocument).filter(QualityDocument.document_no == document_no.strip()).first()
        if item:
            item.category_id = category_id
            item.document_name = document_name.strip()
            item.document_type = (document_type or "").strip() or None
            item.department = (department or "").strip() or None
            item.first_publish_date = first_publish_date
            item.is_active = True
            db.commit()
            db.refresh(item)
            return item
        item = QualityDocument(
            category_id=category_id,
            document_no=document_no.strip(),
            document_name=document_name.strip(),
            document_type=(document_type or "").strip() or None,
            department=(department or "").strip() or None,
            first_publish_date=first_publish_date,
            is_active=True,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def update_document(
    document_id: int,
    category_id: int,
    document_name: str,
    document_type: Optional[str] = None,
    department: Optional[str] = None,
    first_publish_date: Optional[datetime] = None,
    is_active: bool = True,
):
    db = SessionLocal()
    try:
        item = db.query(QualityDocument).filter(QualityDocument.id == document_id).first()
        if not item:
            return None
        item.category_id = category_id
        item.document_name = document_name.strip()
        item.document_type = (document_type or "").strip() or None
        item.department = (department or "").strip() or None
        item.first_publish_date = first_publish_date
        item.is_active = is_active
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def list_document_revisions(document_id: int):
    db = SessionLocal()
    try:
        return (
            db.query(QualityDocumentRevision)
            .filter(QualityDocumentRevision.document_id == document_id)
            .order_by(
                QualityDocumentRevision.published_at.desc().nullslast(),
                QualityDocumentRevision.revision_date.desc().nullslast(),
                QualityDocumentRevision.created_at.desc(),
                QualityDocumentRevision.id.desc(),
            )
            .all()
        )
    finally:
        db.close()


def create_document_revision(
    document_id: int,
    revision_no: str,
    revision_date: Optional[datetime] = None,
    effective_date: Optional[datetime] = None,
    last_review_date: Optional[datetime] = None,
    change_summary: str = "",
    notes: str = "",
    distribution_text: str = "",
    file_path: Optional[str] = None,
):
    db = SessionLocal()
    try:
        item = QualityDocumentRevision(
            document_id=document_id,
            revision_no=revision_no.strip(),
            revision_date=revision_date,
            effective_date=effective_date,
            last_review_date=last_review_date,
            change_summary=change_summary or "",
            notes=notes or "",
            distribution_text=distribution_text or "",
            file_path=(file_path or "").strip() or None,
            status="DRAFT",
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return item
    finally:
        db.close()


def _create_notification_log(
    db,
    document_revision_id: int,
    recipient_email: str,
    recipient_name: str,
    delivery_status: str = "PENDING",
    error_message: str = "",
):
    item = QualityDocumentNotificationLog(
        document_revision_id=document_revision_id,
        recipient_email=recipient_email,
        recipient_name=recipient_name,
        delivery_channel="EMAIL",
        delivery_status=delivery_status,
        sent_at=datetime.utcnow() if delivery_status == "SENT" else None,
        error_message=error_message or "",
    )
    db.add(item)
    return item


def _send_email(to_email: str, subject: str, body: str):
    cfg = _smtp_config()
    if not cfg["enabled"]:
        return False, "QUALITY_SMTP_ENABLED kapalı"
    if not cfg["host"] or not cfg["sender"]:
        return False, "SMTP host/sender eksik"

    message = EmailMessage()
    message["From"] = cfg["sender"]
    message["To"] = to_email
    message["Subject"] = subject
    message.set_content(body)

    try:
        if cfg["use_tls"]:
            context = ssl.create_default_context()
            with smtplib.SMTP(cfg["host"], cfg["port"]) as server:
                server.starttls(context=context)
                if cfg["username"]:
                    server.login(cfg["username"], cfg["password"])
                server.send_message(message)
        else:
            with smtplib.SMTP_SSL(cfg["host"], cfg["port"]) as server:
                if cfg["username"]:
                    server.login(cfg["username"], cfg["password"])
                server.send_message(message)
        return True, ""
    except Exception as exc:
        return False, str(exc)


def publish_document_revision(
    revision_id: int,
    published_by: str = "system",
    software_explanation: str = "",
):
    db = SessionLocal()
    try:
        revision = (
            db.query(QualityDocumentRevision)
            .options(joinedload(QualityDocumentRevision.document))
            .filter(QualityDocumentRevision.id == revision_id)
            .first()
        )
        if not revision:
            return None

        current_published = (
            db.query(QualityDocumentRevision)
            .filter(
                QualityDocumentRevision.document_id == revision.document_id,
                QualityDocumentRevision.status == "PUBLISHED",
                QualityDocumentRevision.id != revision.id,
            )
            .all()
        )
        for old_item in current_published:
            old_item.status = "SUPERSEDED"

        revision.status = "PUBLISHED"
        revision.published_by = published_by
        revision.published_at = datetime.utcnow()
        db.commit()
        db.refresh(revision)

        recipients = _email_candidates_from_distribution_text(revision.distribution_text)
        if not recipients:
            recipients = _fallback_notification_recipients()

        subject = f"Kalite Doküman Revizyonu: {revision.document.document_no} - {revision.document.document_name}"
        body = (
            f"Doküman No: {revision.document.document_no}\n"
            f"Doküman Adı: {revision.document.document_name}\n"
            f"Revizyon No: {revision.revision_no}\n"
            f"Revizyon Tarihi: {_display_date(revision.revision_date)}\n"
            f"Son Kontrol Tarihi: {_display_date(revision.last_review_date)}\n\n"
            f"Revizyon Açıklaması:\n{revision.change_summary or '-'}\n\n"
            f"Personel Bildirim Açıklaması:\n{_normalize_text(software_explanation) or '-'}\n\n"
            f"İç Not:\n{revision.notes or '-'}\n"
        )

        if not recipients:
            _create_notification_log(
                db=db,
                document_revision_id=revision.id,
                recipient_email="",
                recipient_name="Dağıtım listesi bulunamadı",
                delivery_status="PENDING",
                error_message="Dağıtım planında mail adresi yok. QUALITY_NOTIFICATION_EMAILS de tanımlı değil.",
            )
            db.commit()
            return _serialize_revision_entity(revision)

        for recipient in recipients:
            ok, error_message = _send_email(recipient, subject, body)
            _create_notification_log(
                db=db,
                document_revision_id=revision.id,
                recipient_email=recipient,
                recipient_name=recipient,
                delivery_status="SENT" if ok else "PENDING",
                error_message=error_message if not ok else "",
            )
        db.commit()
        return _serialize_revision_entity(revision)
    finally:
        db.close()


def list_notification_logs(document_revision_id: int):
    db = SessionLocal()
    try:
        return (
            db.query(QualityDocumentNotificationLog)
            .filter(QualityDocumentNotificationLog.document_revision_id == document_revision_id)
            .order_by(QualityDocumentNotificationLog.created_at.desc(), QualityDocumentNotificationLog.id.desc())
            .all()
        )
    finally:
        db.close()


def create_and_publish_revision(
    document_id: int,
    revision_no: str,
    revision_date: Optional[datetime] = None,
    effective_date: Optional[datetime] = None,
    last_review_date: Optional[datetime] = None,
    software_explanation: str = "",
    change_summary: str = "",
    notes: str = "",
    distribution_text: str = "",
    file_path: Optional[str] = None,
    published_by: str = "system",
):
    explanation = _normalize_text(software_explanation)
    normalized_revision_no = _normalize_text(revision_no)
    if not normalized_revision_no:
        raise ValueError("Revizyon No zorunludur")
    if not revision_date:
        raise ValueError("Revizyon Tarihi zorunludur")
    if not explanation:
        raise ValueError("Personel Bildirim Açıklaması zorunludur")

    revision = create_document_revision(
        document_id=document_id,
        revision_no=normalized_revision_no,
        revision_date=revision_date,
        effective_date=effective_date,
        last_review_date=last_review_date,
        change_summary=change_summary or "",
        notes=notes,
        distribution_text=distribution_text,
        file_path=file_path,
    )
    return publish_document_revision(
        revision.id,
        published_by=published_by,
        software_explanation=explanation,
    )


def _latest_revision_map(db, document_ids: list[int]) -> dict[int, QualityDocumentRevision]:
    if not document_ids:
        return {}
    revisions = (
        db.query(QualityDocumentRevision)
        .filter(QualityDocumentRevision.document_id.in_(document_ids))
        .order_by(
            QualityDocumentRevision.document_id.asc(),
            QualityDocumentRevision.published_at.desc().nullslast(),
            QualityDocumentRevision.revision_date.desc().nullslast(),
            QualityDocumentRevision.created_at.desc(),
            QualityDocumentRevision.id.desc(),
        )
        .all()
    )
    result: dict[int, QualityDocumentRevision] = {}
    for item in revisions:
        if item.document_id not in result:
            result[item.document_id] = item
    return result


def build_document_row(document: QualityDocument, latest_revision: Optional[QualityDocumentRevision]):
    return {
        "id": document.id,
        "category_id": document.category_id,
        "category_name": document.category.name if document.category else "",
        "DOK. NO": document.document_no,
        "DOKÜMAN ADI": document.document_name,
        "İLK YAY. TAR.": _display_date(document.first_publish_date),
        "REV. NO": latest_revision.revision_no if latest_revision and latest_revision.revision_no else "-",
        "REV. TAR": _display_date(latest_revision.revision_date if latest_revision else None),
        "SON KONTROL TARİHİ": _display_date(latest_revision.last_review_date if latest_revision else None),
        "DOKÜMAN DAĞITIM PLANI": latest_revision.distribution_text if latest_revision and latest_revision.distribution_text else "-",
        "AÇIKLAMALAR": latest_revision.change_summary if latest_revision and latest_revision.change_summary else "-",
        "is_active": document.is_active,
        "latest_revision_status": latest_revision.status if latest_revision else None,
    }


def list_document_rows(search: str = "", active_only: bool = False, category_id: Optional[int] = None):
    db = SessionLocal()
    try:
        query = db.query(QualityDocument).options(joinedload(QualityDocument.category))
        if active_only:
            query = query.filter(QualityDocument.is_active == True)
        if category_id:
            query = query.filter(QualityDocument.category_id == category_id)
        if search.strip():
            token = f"%{search.strip()}%"
            query = query.filter(
                (QualityDocument.document_no.ilike(token)) | (QualityDocument.document_name.ilike(token))
            )
        documents = query.order_by(QualityDocument.document_no.asc(), QualityDocument.id.asc()).all()
        latest_map = _latest_revision_map(db, [doc.id for doc in documents])
        return [build_document_row(doc, latest_map.get(doc.id)) for doc in documents]
    finally:
        db.close()


def get_document_detail(document_id: int):
    db = SessionLocal()
    try:
        document = (
            db.query(QualityDocument)
            .options(joinedload(QualityDocument.category))
            .filter(QualityDocument.id == document_id)
            .first()
        )
        if not document:
            return None
        revisions = (
            db.query(QualityDocumentRevision)
            .filter(QualityDocumentRevision.document_id == document_id)
            .order_by(
                QualityDocumentRevision.published_at.desc().nullslast(),
                QualityDocumentRevision.revision_date.desc().nullslast(),
                QualityDocumentRevision.created_at.desc(),
                QualityDocumentRevision.id.desc(),
            )
            .all()
        )
        notifications = []
        if revisions:
            revision_ids = [item.id for item in revisions]
            notifications = (
                db.query(QualityDocumentNotificationLog)
                .filter(QualityDocumentNotificationLog.document_revision_id.in_(revision_ids))
                .order_by(QualityDocumentNotificationLog.created_at.desc(), QualityDocumentNotificationLog.id.desc())
                .all()
            )
        latest = revisions[0] if revisions else None
        return {
            "document": {
                "id": document.id,
                "category_id": document.category_id,
                "category_name": document.category.name if document.category else "",
                "document_no": document.document_no,
                "document_name": document.document_name,
                "is_active": document.is_active,
            },
            "table_row": build_document_row(document, latest),
            "editor": {
                "category_id": document.category_id,
                "document_no": document.document_no,
                "document_name": document.document_name,
                "first_publish_date": _to_iso(document.first_publish_date),
                "revision_no": latest.revision_no if latest else "",
                "revision_date": _to_iso(latest.revision_date) if latest else None,
                "last_review_date": _to_iso(latest.last_review_date) if latest else None,
                "distribution_text": latest.distribution_text if latest else "",
                "change_summary": latest.change_summary if latest else "",
                "notes": latest.notes if latest else "",
                "is_active": document.is_active,
            },
            "revisions": [
                {
                    "id": item.id,
                    "revision_no": item.revision_no,
                    "revision_date": _to_iso(item.revision_date),
                    "effective_date": _to_iso(item.effective_date),
                    "last_review_date": _to_iso(item.last_review_date),
                    "change_summary": item.change_summary,
                    "notes": item.notes,
                    "distribution_text": item.distribution_text,
                    "status": item.status,
                    "published_by": item.published_by,
                    "published_at": _to_iso(item.published_at),
                }
                for item in revisions
            ],
            "notification_logs": [
                {
                    "id": item.id,
                    "document_revision_id": item.document_revision_id,
                    "recipient_email": item.recipient_email,
                    "recipient_name": item.recipient_name,
                    "delivery_channel": item.delivery_channel,
                    "delivery_status": item.delivery_status,
                    "sent_at": _to_iso(item.sent_at),
                    "error_message": item.error_message,
                    "created_at": _to_iso(item.created_at),
                }
                for item in notifications
            ],
        }
    finally:
        db.close()


def get_quality_overview(search: str = "", active_only: bool = False, category_id: Optional[int] = None):
    categories = list_document_categories(active_only=False)
    rows = list_document_rows(search=search, active_only=active_only, category_id=category_id)
    return {
        "headers": EXCEL_HEADERS,
        "categories": [
            {
                "id": item.id,
                "code": item.code,
                "name": item.name,
                "sort_order": item.sort_order,
                "is_active": item.is_active,
            }
            for item in categories
        ],
        "rows": rows,
        "stats": {
            "total_documents": len(rows),
            "active_documents": sum(1 for item in rows if item.get("is_active")),
            "published_revisions": sum(1 for item in rows if item.get("latest_revision_status") == "PUBLISHED"),
            "draft_or_pending": sum(1 for item in rows if item.get("latest_revision_status") in {None, "DRAFT", "PENDING"}),
        },
    }


def save_document_row(
    *,
    document_id: Optional[int],
    category_id: int,
    document_no: str,
    document_name: str,
    first_publish_date: Optional[datetime],
    revision_no: str,
    revision_date: Optional[datetime],
    last_review_date: Optional[datetime],
    distribution_text: str,
    change_summary: str,
    notes: str,
    is_active: bool,
):
    doc_no = _normalize_text(document_no)
    doc_name = _normalize_text(document_name)
    if not doc_no:
        raise ValueError("DOK. NO zorunludur")
    if not doc_name:
        raise ValueError("DOKÜMAN ADI zorunludur")

    db = SessionLocal()
    try:
        duplicate = db.query(QualityDocument).filter(QualityDocument.document_no == doc_no)
        if document_id:
            duplicate = duplicate.filter(QualityDocument.id != document_id)
        if duplicate.first():
            raise ValueError("Bu DOK. NO zaten kayıtlı")

        if document_id:
            document = db.query(QualityDocument).filter(QualityDocument.id == document_id).first()
            if not document:
                return None
            document.category_id = category_id
            document.document_no = doc_no
            document.document_name = doc_name
            document.first_publish_date = first_publish_date
            document.is_active = is_active
        else:
            document = QualityDocument(
                category_id=category_id,
                document_no=doc_no,
                document_name=doc_name,
                first_publish_date=first_publish_date,
                is_active=is_active,
            )
            db.add(document)
            db.flush()

        latest = (
            db.query(QualityDocumentRevision)
            .filter(QualityDocumentRevision.document_id == document.id)
            .order_by(
                QualityDocumentRevision.published_at.desc().nullslast(),
                QualityDocumentRevision.revision_date.desc().nullslast(),
                QualityDocumentRevision.created_at.desc(),
                QualityDocumentRevision.id.desc(),
            )
            .first()
        )
        if latest:
            latest.revision_no = _normalize_text(revision_no) or "0"
            latest.revision_date = revision_date
            latest.effective_date = revision_date
            latest.last_review_date = last_review_date
            latest.distribution_text = distribution_text or ""
            latest.change_summary = change_summary or ""
            latest.notes = notes or ""
        else:
            latest = QualityDocumentRevision(
                document_id=document.id,
                revision_no=_normalize_text(revision_no) or "0",
                revision_date=revision_date,
                effective_date=revision_date,
                last_review_date=last_review_date,
                distribution_text=distribution_text or "",
                change_summary=change_summary or "",
                notes=notes or "",
                status="DRAFT",
            )
            db.add(latest)

        db.commit()
        db.refresh(document)
        return document
    finally:
        db.close()


def publish_revision_from_row(
    *,
    document_id: int,
    category_id: Optional[int],
    category_name: Optional[str] = None,
    document_no: str,
    document_name: str,
    first_publish_date: Optional[datetime],
    revision_no: str,
    revision_date: Optional[datetime],
    last_review_date: Optional[datetime],
    distribution_text: str = "",
    change_summary: str = "",
    software_explanation: str = "",
    notes: str = "",
    published_by: str = "system",
    is_active: bool = True,
):
    normalized_revision_no = _normalize_text(revision_no)
    if not normalized_revision_no:
        raise ValueError("Revizyon No zorunludur")
    if not revision_date:
        raise ValueError("Revizyon Tarihi zorunludur")
    if not _normalize_text(software_explanation):
        raise ValueError("Personel Bildirim Açıklaması zorunludur")

    db = SessionLocal()
    try:
        current_latest = (
            db.query(QualityDocumentRevision)
            .filter(QualityDocumentRevision.document_id == document_id)
            .order_by(
                QualityDocumentRevision.published_at.desc().nullslast(),
                QualityDocumentRevision.revision_date.desc().nullslast(),
                QualityDocumentRevision.created_at.desc(),
                QualityDocumentRevision.id.desc(),
            )
            .first()
        )
        if current_latest and current_latest.revision_date and revision_date <= current_latest.revision_date:
            raise ValueError("Yeni revizyon tarihi, mevcut revizyon tarihinden daha yeni olmalıdır")
    finally:
        db.close()

    saved = save_document_row(
        document_id=document_id,
        category_id=category_id,
        category_name=category_name,
        document_no=document_no,
        document_name=document_name,
        first_publish_date=first_publish_date,
        revision_no=normalized_revision_no,
        revision_date=revision_date,
        last_review_date=last_review_date,
        distribution_text=distribution_text,
        change_summary=change_summary,
        notes=notes,
        is_active=is_active,
    )
    if not saved:
        return None

    return create_and_publish_revision(
        document_id=document_id,
        revision_no=normalized_revision_no,
        revision_date=revision_date,
        effective_date=revision_date,
        last_review_date=last_review_date,
        software_explanation=software_explanation,
        change_summary=change_summary,
        notes=notes,
        distribution_text=distribution_text,
        published_by=published_by,
    )


def _import_workbook(workbook_path: str):
    wb = load_workbook(workbook_path, data_only=True)
    imported_rows = 0
    updated_rows = 0
    created_documents = 0
    existing_documents = {item.document_no: item for item in list_documents()}

    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        category_code = _normalize_text(sheet_name).upper().replace(" ", "_")
        category = get_or_create_document_category(code=category_code, name=_normalize_text(sheet_name), sort_order=sheet_index * 10)
        header_row_index = None
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = [_normalize_text(cell) for cell in row]
            if "DOK. NO" in normalized and "DOKÜMAN ADI" in normalized:
                header_row_index = row_idx
                headers = {value: idx for idx, value in enumerate(normalized)}
                break
        if not header_row_index:
            continue
        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            dok_no = _normalize_text(row[headers["DOK. NO"]]) if "DOK. NO" in headers else ""
            dok_adi = _normalize_text(row[headers["DOKÜMAN ADI"]]) if "DOKÜMAN ADI" in headers else ""
            if not dok_no or not dok_adi:
                continue
            first_publish_date = _parse_excel_date(row[headers["İLK YAY. TAR."]] if "İLK YAY. TAR." in headers else None)
            revision_no = _normalize_text(row[headers["REV. NO"]] if "REV. NO" in headers else "")
            revision_date = _parse_excel_date(row[headers["REV. TAR"]] if "REV. TAR" in headers else None)
            last_review_date = _parse_excel_date(row[headers["SON KONTROL TARİHİ"]] if "SON KONTROL TARİHİ" in headers else None)
            distribution_text = _normalize_text(row[headers["DOKÜMAN DAĞITIM PLANI"]] if "DOKÜMAN DAĞITIM PLANI" in headers else "")
            notes = _normalize_text(row[headers["AÇIKLAMALAR"]] if "AÇIKLAMALAR" in headers else "")

            existing = existing_documents.get(dok_no)
            document = get_or_create_document(
                category_id=category.id,
                document_no=dok_no,
                document_name=dok_adi,
                first_publish_date=first_publish_date,
            )
            if existing:
                updated_rows += 1
            else:
                created_documents += 1
                existing_documents[dok_no] = document

            save_document_row(
                document_id=document.id,
                category_id=category.id,
                document_no=dok_no,
                document_name=dok_adi,
                first_publish_date=first_publish_date,
                revision_no=revision_no or "0",
                revision_date=revision_date,
                last_review_date=last_review_date,
                distribution_text=distribution_text,
                change_summary=notes,
                notes=notes,
                is_active=True,
            )
            imported_rows += 1
    return {
        "ok": True,
        "imported_rows": imported_rows,
        "updated_rows": updated_rows,
        "created_documents": created_documents,
    }


def import_lt03_file(file_path: str):
    return _import_workbook(file_path)


def import_lt03_upload(file_bytes: bytes, filename: str):
    suffix = Path(filename or "lt03.xlsx").suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
        temp.write(file_bytes)
        temp_path = temp.name
    try:
        return _import_workbook(temp_path)
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass
