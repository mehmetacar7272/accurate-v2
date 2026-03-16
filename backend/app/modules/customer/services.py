from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.modules.customer.models import Customer, CustomerBranch, CustomerContact

REQUIRED_HEADERS = [
    'Müşteri Adı', 'Ticari Unvan', 'Yetkili Kişi', 'Vergi D.', 'Vergi No', 'Telefon', 'E-Posta', 'Adres'
]


def _serialize_contact(item: CustomerContact):
    return {
        'id': item.id,
        'branch_id': item.branch_id,
        'full_name': item.full_name,
        'phone': item.phone,
        'email': item.email,
        'title': item.title,
        'is_default': item.is_default,
        'is_active': item.is_active,
    }


def _serialize_branch(item: CustomerBranch):
    return {
        'id': item.id,
        'branch_name': item.branch_name,
        'address': item.address,
        'is_default': item.is_default,
        'is_active': item.is_active,
    }


def _serialize_customer(item: Customer):
    return {
        'id': item.id,
        'customer_name': item.customer_name,
        'trade_name': item.trade_name,
        'tax_office': item.tax_office,
        'tax_number': item.tax_number,
        'is_active': item.is_active,
        'branches': [_serialize_branch(x) for x in sorted(item.branches, key=lambda b: (not b.is_default, b.id)) if x.is_active],
        'contacts': [_serialize_contact(x) for x in sorted(item.contacts, key=lambda c: (not c.is_default, c.id)) if x.is_active],
    }


def _normalize_text(value: Any) -> str:
    return str(value or '').strip()


def _build_branch_rows(data: dict) -> list[dict[str, Any]]:
    rows = []
    provided = data.get('branches') or []
    for index, item in enumerate(provided):
        branch_name = _normalize_text(item.get('branch_name'))
        address = _normalize_text(item.get('address'))
        if not branch_name and not address:
            continue
        rows.append({
            'id': item.get('id'),
            'client_key': _normalize_text(item.get('client_key')) or f'branch-{index + 1}',
            'branch_name': branch_name or 'Merkez',
            'address': address,
            'is_default': bool(item.get('is_default')),
        })

    if rows:
        if not any(row['is_default'] for row in rows):
            rows[0]['is_default'] = True
        return rows

    branch_name = _normalize_text(data.get('branch_name')) or 'Merkez'
    address = _normalize_text(data.get('address'))
    if not address:
        raise ValueError('Zorunlu alanlar: Adres')
    return [{
        'id': None,
        'client_key': 'branch-1',
        'branch_name': branch_name,
        'address': address,
        'is_default': True,
    }]


def _build_contact_rows(data: dict, branches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    provided = data.get('contacts') or []
    default_branch_key = next((row['client_key'] for row in branches if row['is_default']), branches[0]['client_key'])

    for index, item in enumerate(provided):
        full_name = _normalize_text(item.get('full_name'))
        phone = _normalize_text(item.get('phone'))
        email = _normalize_text(item.get('email'))
        title = _normalize_text(item.get('title'))
        if not full_name and not phone and not email and not title:
            continue
        rows.append({
            'id': item.get('id'),
            'client_key': _normalize_text(item.get('client_key')) or f'contact-{index + 1}',
            'branch_key': _normalize_text(item.get('branch_key')) or default_branch_key,
            'full_name': full_name,
            'phone': phone,
            'email': email,
            'title': title,
            'is_default': bool(item.get('is_default')),
        })

    if rows:
        missing = []
        for row in rows:
            required = []
            if not row['full_name']:
                required.append('Yetkili Kişi')
            if not row['phone']:
                required.append('Telefon')
            if not row['email']:
                required.append('E-Posta')
            if not row['title']:
                required.append('Görev')
            if required:
                missing.extend(required)
        if missing:
            deduped = []
            for item in missing:
                if item not in deduped:
                    deduped.append(item)
            raise ValueError('Zorunlu alanlar: ' + ', '.join(deduped))
        if not any(row['is_default'] for row in rows):
            rows[0]['is_default'] = True
        return rows

    full_name = _normalize_text(data.get('contact_name'))
    phone = _normalize_text(data.get('phone'))
    email = _normalize_text(data.get('email'))
    title = _normalize_text(data.get('title'))
    missing = []
    if not full_name:
        missing.append('Yetkili Kişi')
    if not phone:
        missing.append('Telefon')
    if not email:
        missing.append('E-Posta')
    if not title:
        missing.append('Görev')
    if missing:
        raise ValueError('Zorunlu alanlar: ' + ', '.join(missing))
    return [{
        'id': None,
        'client_key': 'contact-1',
        'branch_key': default_branch_key,
        'full_name': full_name,
        'phone': phone,
        'email': email,
        'title': title,
        'is_default': True,
    }]


def list_customers(query: str | None = None):
    db = SessionLocal()
    try:
        items = db.query(Customer).filter(Customer.is_active == True).order_by(Customer.trade_name.asc(), Customer.id.asc()).all()
        if query and query.strip():
            needle = query.strip().lower()
            filtered = []
            for item in items:
                haystack = ' '.join([
                    _normalize_text(item.customer_name),
                    _normalize_text(item.trade_name),
                    _normalize_text(item.tax_number),
                    ' '.join(_normalize_text(branch.branch_name) + ' ' + _normalize_text(branch.address) for branch in item.branches if branch.is_active),
                    ' '.join(
                        ' '.join([
                            _normalize_text(contact.full_name),
                            _normalize_text(contact.phone),
                            _normalize_text(contact.email),
                            _normalize_text(contact.title),
                        ])
                        for contact in item.contacts if contact.is_active
                    ),
                ]).lower()
                if needle in haystack:
                    filtered.append(item)
            items = filtered
        return [_serialize_customer(x) for x in items]
    finally:
        db.close()


def _validate_customer_header(data: dict):
    customer_name = _normalize_text(data.get('customer_name'))
    trade_name = _normalize_text(data.get('trade_name'))
    tax_office = _normalize_text(data.get('tax_office'))
    tax_number = _normalize_text(data.get('tax_number'))

    missing = []
    if not customer_name:
        missing.append('Müşteri Adı')
    if not trade_name:
        missing.append('Ticari Unvan')
    if not tax_office:
        missing.append('Vergi Dairesi')
    if not tax_number:
        missing.append('Vergi No')
    if missing:
        raise ValueError('Zorunlu alanlar: ' + ', '.join(missing))

    return customer_name, trade_name, tax_office, tax_number


def create_customer(data: dict):
    db = SessionLocal()
    try:
        customer_name, trade_name, tax_office, tax_number = _validate_customer_header(data)
        branch_rows = _build_branch_rows(data)
        contact_rows = _build_contact_rows(data, branch_rows)

        customer = db.query(Customer).filter(Customer.trade_name == trade_name, Customer.tax_number == tax_number).first()
        if customer:
            raise ValueError('Aynı ticari unvan ve vergi numarası ile kayıtlı müşteri zaten var')

        customer = Customer(
            customer_name=customer_name,
            trade_name=trade_name,
            tax_office=tax_office,
            tax_number=tax_number,
            is_active=True,
        )
        db.add(customer)
        db.flush()

        branch_map: dict[str, CustomerBranch] = {}
        for row in branch_rows:
            branch = CustomerBranch(
                customer_id=customer.id,
                branch_name=row['branch_name'],
                address=row['address'],
                is_default=False,
                is_active=True,
            )
            db.add(branch)
            db.flush()
            branch_map[row['client_key']] = branch

        default_branch_key = next((row['client_key'] for row in branch_rows if row['is_default']), branch_rows[0]['client_key'])
        default_branch_id = branch_map[default_branch_key].id

        for branch in branch_map.values():
            branch.is_default = branch.id == default_branch_id

        contact_map: dict[str, CustomerContact] = {}
        for row in contact_rows:
            branch = branch_map.get(row['branch_key']) or branch_map[default_branch_key]
            contact = CustomerContact(
                customer_id=customer.id,
                branch_id=branch.id,
                full_name=row['full_name'],
                phone=row['phone'],
                email=row['email'],
                title=row['title'],
                is_default=False,
                is_active=True,
            )
            db.add(contact)
            db.flush()
            contact_map[row['client_key']] = contact

        default_contact_key = next((row['client_key'] for row in contact_rows if row['is_default']), contact_rows[0]['client_key'])
        default_contact_id = contact_map[default_contact_key].id

        for contact in contact_map.values():
            contact.is_default = contact.id == default_contact_id

        db.commit()
        db.refresh(customer)
        return _serialize_customer(customer)
    finally:
        db.close()


def update_customer(customer_id: int, data: dict):
    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id, Customer.is_active == True).first()
        if not customer:
            raise ValueError('Müşteri bulunamadı')

        customer_name, trade_name, tax_office, tax_number = _validate_customer_header(data)
        branch_rows = _build_branch_rows(data)
        contact_rows = _build_contact_rows(data, branch_rows)

        duplicate = db.query(Customer).filter(
            Customer.id != customer_id,
            Customer.trade_name == trade_name,
            Customer.tax_number == tax_number,
            Customer.is_active == True,
        ).first()
        if duplicate:
            raise ValueError('Aynı ticari unvan ve vergi numarası ile kayıtlı başka müşteri var')

        customer.customer_name = customer_name
        customer.trade_name = trade_name
        customer.tax_office = tax_office
        customer.tax_number = tax_number
        customer.is_active = True

        existing_branches = db.query(CustomerBranch).filter(CustomerBranch.customer_id == customer.id).all()
        existing_contacts = db.query(CustomerContact).filter(CustomerContact.customer_id == customer.id).all()

        branch_by_id = {item.id: item for item in existing_branches}
        contact_by_id = {item.id: item for item in existing_contacts}

        active_branch_ids = set()
        branch_map: dict[str, CustomerBranch] = {}

        for row in branch_rows:
            branch = None
            row_id = row.get('id')
            if row_id:
                branch = branch_by_id.get(row_id)
            if not branch:
                branch = CustomerBranch(
                    customer_id=customer.id,
                    branch_name=row['branch_name'],
                    address=row['address'],
                    is_default=False,
                    is_active=True,
                )
                db.add(branch)
                db.flush()
            branch.branch_name = row['branch_name']
            branch.address = row['address']
            branch.is_active = True
            active_branch_ids.add(branch.id)
            branch_map[row['client_key']] = branch

        for branch in existing_branches:
            if branch.id not in active_branch_ids:
                branch.is_active = False
                branch.is_default = False

        default_branch_key = next((row['client_key'] for row in branch_rows if row['is_default']), branch_rows[0]['client_key'])
        default_branch = branch_map[default_branch_key]
        for branch in existing_branches + [item for item in branch_map.values() if item.id not in {b.id for b in existing_branches}]:
            if branch.id in active_branch_ids:
                branch.is_default = branch.id == default_branch.id

        active_contact_ids = set()
        contact_map: dict[str, CustomerContact] = {}

        for row in contact_rows:
            contact = None
            row_id = row.get('id')
            if row_id:
                contact = contact_by_id.get(row_id)
            branch = branch_map.get(row['branch_key']) or default_branch
            if not contact:
                contact = CustomerContact(
                    customer_id=customer.id,
                    branch_id=branch.id,
                    full_name=row['full_name'],
                    phone=row['phone'],
                    email=row['email'],
                    title=row['title'],
                    is_default=False,
                    is_active=True,
                )
                db.add(contact)
                db.flush()
            contact.branch_id = branch.id
            contact.full_name = row['full_name']
            contact.phone = row['phone']
            contact.email = row['email']
            contact.title = row['title']
            contact.is_active = True
            active_contact_ids.add(contact.id)
            contact_map[row['client_key']] = contact

        for contact in existing_contacts:
            if contact.id not in active_contact_ids:
                contact.is_active = False
                contact.is_default = False

        default_contact_key = next((row['client_key'] for row in contact_rows if row['is_default']), contact_rows[0]['client_key'])
        default_contact = contact_map[default_contact_key]
        for contact in existing_contacts + [item for item in contact_map.values() if item.id not in {c.id for c in existing_contacts}]:
            if contact.id in active_contact_ids:
                contact.is_default = contact.id == default_contact.id

        db.commit()
        db.refresh(customer)
        return _serialize_customer(customer)
    finally:
        db.close()


def deactivate_customer(customer_id: int):
    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id, Customer.is_active == True).first()
        if not customer:
            raise ValueError('Müşteri bulunamadı')

        customer.is_active = False
        for branch in customer.branches:
            branch.is_active = False
            branch.is_default = False
        for contact in customer.contacts:
            contact.is_active = False
            contact.is_default = False

        db.commit()
        return {'ok': True, 'customer_id': customer_id}
    finally:
        db.close()


def import_customers_from_excel(content: bytes):
    workbook = load_workbook(filename=BytesIO(content))
    sheet = workbook.active

    headers = [str(cell.value or '').strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers[:len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        raise ValueError('Excel başlıkları beklenen formatta değil')

    imported_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        create_customer({
            'customer_name': _normalize_text(row[0]),
            'trade_name': _normalize_text(row[1]),
            'contact_name': _normalize_text(row[2]),
            'tax_office': _normalize_text(row[3]),
            'tax_number': _normalize_text(row[4]),
            'phone': _normalize_text(row[5]),
            'email': _normalize_text(row[6]),
            'address': _normalize_text(row[7]),
            'title': 'Yetkili',
            'branch_name': 'Merkez',
        })
        imported_rows += 1

    return {'ok': True, 'imported_rows': imported_rows}
